from __future__ import annotations

import logging
import re
import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from src.models import GptScoreResult, ProjectFull

logger = logging.getLogger(__name__)

JOURNAL_COLUMNS = [
    "№",
    "Дата отклика",
    "Площадка",
    "Ссылка на проект",
    "Тип проекта",
    "Статус",
    "Результат общения",
    "Предложения",
    "Отклик",
]

OFFER_COLUMN = 8
RESPONSE_COLUMN = 9
TEXT_COLUMN_WIDTH = 73

PLATFORM_DISPLAY = {
    "kwork": "Kwork",
    "flru": "FL.ru",
    "telegram": "Telegram",
}


def kwork_project_url(project_id: str) -> str:
    return f"https://kwork.ru/projects/{project_id}/view"


def infer_project_type(title: str) -> str:
    low = (title or "").lower()
    if any(x in low for x in ("telegram", "телеграм", "бот")):
        return "Telegram-бот"
    if "парс" in low or "parser" in low or "scrap" in low:
        return "Парсинг"
    if any(x in low for x in ("rag", "llm", "gpt", "ии ", " ai", "нейро")):
        return "AI/RAG"
    if any(x in low for x in ("сайт", "лендинг", "wordpress", "веб", "frontend", "backend")):
        return "Веб-MVP"
    if any(x in low for x in ("интеграц", "api", "crm", "1с")):
        return "Интеграция"
    if any(x in low for x in ("автомат", "скрипт", "excel", "google sheet")):
        return "Автоматизация"
    return "Другое"


def format_response_payload(
    response_text: str,
    *,
    price: str | None = None,
    delivery_days: int | None = None,
) -> str:
    price_text = f"{price} ₽" if price else "—"
    days_text = f"{delivery_days} дн." if delivery_days else "—"
    body = (response_text or "").strip()
    return f"Отклик\n{body}\n\nОбщий бюджет: {price_text}\nСрок: {days_text}".strip()


class JournalWriter:
    def __init__(self, journal_path: str | Path) -> None:
        self.journal_path = Path(journal_path)

    @staticmethod
    def _url_from_cell(cell) -> str:
        link = getattr(cell, "hyperlink", None)
        if link is not None and getattr(link, "target", None):
            return str(link.target)
        return str(cell.value or "")

    @staticmethod
    def _set_url_cell(ws, row: int, url: str) -> None:
        cell = ws.cell(row=row, column=4, value=url)
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")

    def _ensure_workbook(self) -> None:
        if not self.journal_path.exists():
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Отклики"
            for col, header in enumerate(JOURNAL_COLUMNS, start=1):
                ws.cell(row=1, column=col, value=header)
            self._apply_journal_layout(ws)
            self.journal_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(self.journal_path)

    def _apply_journal_layout(self, ws) -> None:
        header = self._header_row(ws)
        ws.column_dimensions["H"].width = TEXT_COLUMN_WIDTH
        ws.column_dimensions["I"].width = TEXT_COLUMN_WIDTH

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.cell(row=header, column=OFFER_COLUMN, value="Предложения")
        ws.cell(row=header, column=RESPONSE_COLUMN, value="Отклик")
        for col in (OFFER_COLUMN, RESPONSE_COLUMN):
            cell = ws.cell(row=header, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        text_align = Alignment(vertical="top", wrap_text=True)
        for row in range(header + 1, ws.max_row + 1):
            for col in (OFFER_COLUMN, RESPONSE_COLUMN):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.alignment = text_align

    def normalize_layout(self) -> bool:
        if not self.journal_path.exists():
            return False
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        wb.save(self.journal_path)
        wb.close()
        return True

    def _header_row(self, ws) -> int:
        for row in range(1, min(ws.max_row + 1, 20)):
            if ws.cell(row=row, column=2).value == "Дата отклика":
                return row
        return 1

    def _writable(self, ws, row: int, col: int) -> bool:
        return not isinstance(ws.cell(row=row, column=col), MergedCell)

    def _row_has_data(self, ws, row: int) -> bool:
        for col in (2, 4):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                return True
        return False

    def _next_row(self, ws) -> int:
        header = self._header_row(ws)
        for row in range(header + 1, max(ws.max_row + 2, header + 250)):
            if not self._writable(ws, row, 2):
                continue
            if not self._row_has_data(ws, row):
                return row
        return ws.max_row + 1

    def _next_number(self, ws) -> int:
        header = self._header_row(ws)
        max_n = 0
        for row in range(header + 1, ws.max_row + 1):
            if not self._row_has_data(ws, row):
                continue
            cell = ws.cell(row=row, column=1)
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value is not None:
                    max_n = max(max_n, int(cell.value))
            except (TypeError, ValueError):
                continue
        return max_n + 1

    def project_ids_in_journal(self) -> set[str]:
        if not self.journal_path.exists():
            return set()
        wb = load_workbook(self.journal_path, read_only=True, data_only=True)
        ws = wb.active
        ids: set[str] = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
            cell = row[0]
            if cell.value is None:
                continue
            text = self._url_from_cell(cell)
            match = re.search(r"/projects/(\d+)", text.replace("\\", "/"))
            if match:
                ids.add(match.group(1))
                continue
            for part in text.replace("\\", "/").split("/"):
                if part.isdigit() and len(part) >= 3:
                    ids.add(part)
        wb.close()
        return ids

    def _find_row_by_project_id(self, ws, project_id: str) -> int | None:
        header = self._header_row(ws)
        for row in range(header + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=4)
            if isinstance(cell, MergedCell):
                continue
            if project_id in self._url_from_cell(cell):
                return row
        return None

    def update_status_by_project_id(
        self,
        project_id: str,
        *,
        status: str,
        result: str | None = None,
    ) -> bool:
        if not self.journal_path.exists():
            return False
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._find_row_by_project_id(ws, project_id)
        if row is None:
            wb.close()
            return False
        changed = False
        if self._writable(ws, row, 6):
            current = str(ws.cell(row=row, column=6).value or "").strip()
            if current != status:
                ws.cell(row=row, column=6, value=status)
                changed = True
        if result is not None and self._writable(ws, row, 7):
            current_result = str(ws.cell(row=row, column=7).value or "").strip()
            if current_result != result:
                ws.cell(row=row, column=7, value=result)
                changed = True
        if changed:
            wb.save(self.journal_path)
            logger.info(
                "journal_status_updated project_id=%s row=%d status=%r",
                project_id,
                row,
                status,
            )
        wb.close()
        return changed

    def update_response_by_project_id(
        self,
        project_id: str,
        response_payload: str,
    ) -> bool:
        if not self.journal_path.exists():
            return False
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._find_row_by_project_id(ws, project_id)
        if row is None or not self._writable(ws, row, 9):
            wb.close()
            return False
        ws.cell(row=row, column=9, value=response_payload)
        ws.cell(row=row, column=9).alignment = Alignment(vertical="top", wrap_text=True)
        wb.save(self.journal_path)
        wb.close()
        logger.info("journal_response_updated project_id=%s row=%d", project_id, row)
        return True

    def update_notes_by_project_id(self, project_id: str, notes: str) -> bool:
        return self.update_response_by_project_id(project_id, notes)

    def remove_row_by_project_id(self, project_id: str) -> bool:
        if not self.journal_path.exists():
            return False
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._find_row_by_project_id(ws, project_id)
        if row is None:
            wb.close()
            return False
        for col in range(1, len(JOURNAL_COLUMNS) + 1):
            if self._writable(ws, row, col):
                ws.cell(row=row, column=col, value=None)
        wb.save(self.journal_path)
        wb.close()
        logger.info("journal_row_removed project_id=%s row=%d", project_id, row)
        return True

    def _repair_row_cells(
        self,
        ws,
        row: int,
        project_id: str,
        *,
        title: str = "",
        project_type: str | None = None,
    ) -> bool:
        changed = False
        url_text = self._url_from_cell(ws.cell(row=row, column=4)).strip()
        if url_text.startswith("http"):
            url_cell = ws.cell(row=row, column=4)
            link = getattr(url_cell, "hyperlink", None)
            if link is None or not getattr(link, "target", None):
                self._set_url_cell(ws, row, url_text)
                changed = True

        type_val = str(ws.cell(row=row, column=5).value or "").strip()
        if type_val in ("", "—", "-"):
            offer_text = str(ws.cell(row=row, column=8).value or "")
            infer_title = title or offer_text.split("\n")[0]
            new_type = (project_type or "").strip() or infer_project_type(infer_title)
            ws.cell(row=row, column=5, value=new_type)
            changed = True
        return changed

    def repair_row_by_project_id(
        self,
        project_id: str,
        *,
        title: str = "",
        project_type: str | None = None,
    ) -> bool:
        if not self.journal_path.exists():
            return False
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._find_row_by_project_id(ws, project_id)
        if row is None:
            wb.close()
            return False
        changed = self._repair_row_cells(
            ws,
            row,
            project_id,
            title=title,
            project_type=project_type,
        )
        if changed:
            wb.save(self.journal_path)
        wb.close()
        return changed

    def repair_all_rows(
        self,
        *,
        titles: dict[str, str] | None = None,
        project_types: dict[str, str] | None = None,
    ) -> int:
        if not self.journal_path.exists():
            return 0
        titles = titles or {}
        project_types = project_types or {}
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        changed_rows = 0
        for row in range(2, ws.max_row + 1):
            if not self._writable(ws, row, 4):
                continue
            url_text = self._url_from_cell(ws.cell(row=row, column=4))
            match = re.search(r"/projects/(\d+)", url_text.replace("\\", "/"))
            if not match:
                continue
            project_id = match.group(1)
            if self._repair_row_cells(
                ws,
                row,
                project_id,
                title=titles.get(project_id, ""),
                project_type=project_types.get(project_id),
            ):
                changed_rows += 1
        if changed_rows:
            wb.save(self.journal_path)
        wb.close()
        return changed_rows

    def append_submission(
        self,
        project: ProjectFull,
        score: GptScoreResult,
        response_text: str,
    ) -> int:
        self._ensure_workbook()
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._next_row(ws)

        platform_label = PLATFORM_DISPLAY.get(project.platform, project.platform)
        response_payload = format_response_payload(
            response_text,
            price=project.desired_budget or project.max_budget,
        )

        ws.cell(row=row, column=1, value=self._next_number(ws))
        ws.cell(row=row, column=2, value=date.today())
        ws.cell(row=row, column=3, value=platform_label)
        self._set_url_cell(ws, row, project.url)
        ws.cell(row=row, column=5, value=score.suggested_project_type)
        ws.cell(row=row, column=6, value="Отправлен")
        ws.cell(row=row, column=7, value="Жду ответа")
        ws.cell(row=row, column=8, value=project.full_description)
        ws.cell(row=row, column=9, value=response_payload)
        ws.cell(row=row, column=8).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=9).alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(self.journal_path)
        logger.info(
            "journal_appended project_id=%s row=%d",
            project.project_id,
            row,
        )
        return row

    def append_prepared(
        self,
        project: ProjectFull,
        score: GptScoreResult,
        response_text: str,
        *,
        price: str | None = None,
        delivery_days: int | None = None,
    ) -> int:
        self._ensure_workbook()
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._next_row(ws)

        platform_label = PLATFORM_DISPLAY.get(project.platform, project.platform)
        response_payload = format_response_payload(
            response_text,
            price=price,
            delivery_days=delivery_days,
        )

        ws.cell(row=row, column=1, value=self._next_number(ws))
        ws.cell(row=row, column=2, value=date.today())
        ws.cell(row=row, column=3, value=platform_label)
        self._set_url_cell(ws, row, project.url)
        ws.cell(row=row, column=5, value=score.suggested_project_type)
        ws.cell(row=row, column=6, value="Подготовлен")
        ws.cell(row=row, column=7, value="Жду ответа")
        ws.cell(row=row, column=8, value=project.full_description)
        ws.cell(row=row, column=9, value=response_payload)
        ws.cell(row=row, column=8).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=9).alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(self.journal_path)
        logger.info(
            "journal_prepared project_id=%s row=%d",
            project.project_id,
            row,
        )
        return row

    def append_kwork_offer_status(
        self,
        *,
        project_id: str,
        title: str,
        status: str,
        result: str,
        project_type: str | None = None,
    ) -> int:
        self._ensure_workbook()
        wb = load_workbook(self.journal_path)
        ws = wb.active
        self._apply_journal_layout(ws)
        row = self._next_row(ws)
        url = kwork_project_url(project_id)
        type_label = (project_type or "").strip() or infer_project_type(title)

        ws.cell(row=row, column=1, value=self._next_number(ws))
        ws.cell(row=row, column=2, value=date.today())
        ws.cell(row=row, column=3, value="Kwork")
        self._set_url_cell(ws, row, url)
        ws.cell(row=row, column=5, value=type_label)
        ws.cell(row=row, column=6, value=status)
        ws.cell(row=row, column=7, value=result)
        ws.cell(row=row, column=8, value=title.strip())
        ws.cell(row=row, column=9, value="")
        ws.cell(row=row, column=8).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=9).alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(self.journal_path)
        wb.close()
        logger.info(
            "journal_offer_synced project_id=%s row=%d status=%r",
            project_id,
            row,
            status,
        )
        return row

    @staticmethod
    def create_template_copy(dest: Path, source: Path | None = None) -> Path:
        dest.parent.mkdir(parents=True, exist_ok=True)
        if source and source.exists():
            shutil.copy(source, dest)
        else:
            JournalWriter(dest)._ensure_workbook()
        return dest
