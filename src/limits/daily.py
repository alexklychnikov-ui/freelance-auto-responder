from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def count_today_responses(journal_path: str | Path) -> int:
    path = Path(journal_path)
    if not path.exists():
        return 0

    today = date.today().isoformat()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        if str(row[1]) == today:
            count += 1
    wb.close()
    return count


def is_daily_limit_reached(journal_path: str | Path, max_daily: int) -> bool:
    return count_today_responses(journal_path) >= max_daily


def count_today_platform_prepared(prepared_store: Any, platform: str) -> int:
    """Soft daily awareness: prepared/confirmed drafts today for a platform."""
    today = date.today()
    count = 0
    for item in prepared_store.list_all():
        if getattr(item, "platform", None) != platform:
            continue
        prepared_at = getattr(item, "prepared_at", None)
        if not isinstance(prepared_at, datetime):
            continue
        day = (
            prepared_at.astimezone(timezone.utc).date()
            if prepared_at.tzinfo
            else prepared_at.date()
        )
        if day == today:
            count += 1
    return count
