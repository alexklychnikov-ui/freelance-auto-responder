from openpyxl import load_workbook

path = "/opt/freelance-responder/data/response_journal.xlsx"
wb = load_workbook(path)
ws = wb.active
print("sheet:", ws.title, "max_row:", ws.max_row)
for row in range(1, min(ws.max_row + 1, 15)):
    vals = [ws.cell(row=row, column=c).value for c in range(1, 9)]
    print(row, vals)
