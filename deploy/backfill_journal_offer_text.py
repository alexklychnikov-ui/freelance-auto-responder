from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from src.adapters.kwork import KworkAdapter
from src.adapters.kwork_auth import KworkCredentials
from src.browser.factory import close_browser_client, get_browser_client
from src.config import Settings, get_settings, load_sources
from src.journal.writer import JournalWriter


def _extract_project_id(url: str) -> str | None:
    match = re.search(r"/projects/(\d+)", (url or "").replace("\\", "/"))
    if match:
        return match.group(1)
    return None


def _build_kwork_adapter(settings: Settings) -> KworkAdapter:
    sources = [s for s in load_sources(settings.sources_config_path) if s.platform == "kwork"]
    source = next((s for s in sources if s.enabled), None) or (sources[0] if sources else None)
    if source is None:
        raise RuntimeError("Kwork source not found in config/sources.yaml")

    browser = get_browser_client(settings)
    creds_raw = settings.kwork_credentials()
    creds = KworkCredentials(*creds_raw) if creds_raw else None
    return KworkAdapter(
        source_key=source.id,
        listing_url=source.url or "https://kwork.ru/projects?c=11",
        browser=browser,
        dry_run_submit=settings.dry_run_submit,
        kwork_credentials=creds,
        auto_login=settings.kwork_auto_login,
    )


def backfill_offer_text(
    journal_path: Path, *, limit: int | None = None, verbose: bool = False
) -> tuple[int, int, int]:
    settings = get_settings()
    adapter = _build_kwork_adapter(settings)
    browser = adapter.browser
    updated = 0
    skipped = 0
    failed = 0

    try:
        writer = JournalWriter(journal_path)
        writer.normalize_layout()

        wb = load_workbook(journal_path)
        ws = wb.active
        header = writer._header_row(ws)  # noqa: SLF001
        processed = 0

        for row in range(header + 1, ws.max_row + 1):
            if limit is not None and processed >= limit:
                break
            if not writer._writable(ws, row, 4):  # noqa: SLF001
                continue

            url = writer._url_from_cell(ws.cell(row=row, column=4)).strip()  # noqa: SLF001
            project_id = _extract_project_id(url)
            if not project_id:
                skipped += 1
                continue

            try:
                project = adapter.read_full(project_id)
                offer_text = (project.full_description or "").strip()
                if not offer_text:
                    failed += 1
                    continue
                ws.cell(row=row, column=8, value=offer_text)
                ws.cell(row=row, column=8).alignment = Alignment(vertical="top", wrap_text=True)
                updated += 1
            except Exception as exc:
                if verbose:
                    print(f"FAIL row={row} project_id={project_id}: {exc}")
                failed += 1
            finally:
                processed += 1

        wb.save(journal_path)
        wb.close()
    finally:
        close_browser_client(browser)

    return updated, skipped, failed


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Backfill column H (Предложения) from Kwork project cards."
    )
    parser.add_argument("--journal-path", required=True, help="Path to response_journal.xlsx")
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Max number of rows to process (for dry run batches)",
    )
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    journal_path = Path(args.journal_path)
    if not journal_path.exists():
        print(f"FAIL: journal not found: {journal_path}")
        return 1

    updated, skipped, failed = backfill_offer_text(
        journal_path, limit=args.limit, verbose=args.verbose
    )
    print(f"DONE: updated={updated} skipped={skipped} failed={failed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

