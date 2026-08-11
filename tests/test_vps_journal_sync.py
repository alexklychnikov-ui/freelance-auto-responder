from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import load_workbook

from src.adapters.kwork import OfferFormSnapshot
from src.adapters.kwork_offers import KworkOfferComment
from src.journal.kwork_status_sync import KworkStatusSyncResult
from src.journal.vps_sync import sync_journal_on_vps
from src.journal.writer import JournalWriter, format_response_payload
from src.models import GptScoreResult, ProjectFull
from src.responses.prepared_store import PreparedResponse, PreparedResponseStore


def _prepared(project_id: str, *, exported: bool = False) -> PreparedResponse:
    project = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id=project_id,
        url=f"https://kwork.ru/projects/{project_id}",
        title=f"Project {project_id}",
        full_description="desc",
    )
    score = GptScoreResult(
        score=8,
        fit=True,
        reason="ok",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Telegram-бот",
        competition_level="low",
        recommendation="откликаться",
    )
    return PreparedResponse(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id=project_id,
        url=project.url,
        title=project.title,
        project=project,
        score=score,
        response_text="AI DRAFT TEXT",
        price="1000",
        delivery_days=7,
        prepared_at=datetime.now(timezone.utc),
        journal_confirmed=True,
        journal_exported=exported,
    )


def test_sync_journal_on_vps_appends_and_marks_exported(tmp_path, monkeypatch) -> None:
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    store = PreparedResponseStore(tmp_path / "prepared")
    item = _prepared("3204427")
    store.save(item)

    def _fake_get_browser(settings):
        raise ModuleNotFoundError("mcp")

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        _fake_get_browser,
    )

    def _fake_offers(*args, **kwargs):
        return KworkStatusSyncResult(updated=1, matched=1, appended=2)

    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        _fake_offers,
    )
    settings = type("S", (), {"response_journal": str(journal_path)})()
    result = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)

    assert result.appended_prepared == 1
    assert result.offers_updated == 1
    assert result.offers_appended == 2
    saved = store.load("kwork", "kwork_dev_it", "3204427")
    assert saved is not None and saved.journal_exported is True


def test_sync_does_not_overwrite_existing_i_with_ai_when_kwork_fails(
    tmp_path, monkeypatch
) -> None:
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    item = _prepared("3217871", exported=True)
    store = PreparedResponseStore(tmp_path / "prepared")
    store.save(item)

    original = format_response_payload(
        "PLATFORM TEXT ALREADY IN JOURNAL", price="9000", delivery_days=10
    )
    writer.append_prepared(
        item.project,
        item.score,
        "PLATFORM TEXT ALREADY IN JOURNAL",
        price="9000",
        delivery_days=10,
    )
    writer.update_response_by_project_id("3217871", original)

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        lambda settings: object(),
    )
    monkeypatch.setattr(
        "src.browser.factory.close_browser_client",
        lambda client: None,
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        lambda *a, **k: KworkStatusSyncResult(),
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.fetch_my_offer_comment_details",
        lambda browser, navigate=False: {},
    )
    monkeypatch.setattr(
        "src.journal.vps_sync._read_offer_text_from_new_offer_form",
        lambda browser, pid: OfferFormSnapshot(
            description="", ok=False, error="form_missing"
        ),
    )

    settings = type("S", (), {"response_journal": str(journal_path)})()
    result = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)

    assert result.updated_notes == 0
    wb = load_workbook(journal_path)
    cell = str(wb.active.cell(row=2, column=9).value or "")
    assert "PLATFORM TEXT ALREADY IN JOURNAL" in cell
    assert "AI DRAFT TEXT" not in cell


def _prepared_telegram_tz(project_id: str, *, exported: bool = False) -> PreparedResponse:
    project = ProjectFull(
        platform="telegram",
        source_key="tz_manual",
        project_id=project_id,
        url="",
        title=f"TZ {project_id}",
        full_description="tz desc",
    )
    score = GptScoreResult(
        score=8,
        fit=True,
        reason="ok",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Telegram-бот",
        competition_level="low",
        recommendation="откликаться",
    )
    return PreparedResponse(
        platform="telegram",
        source_key="tz_manual",
        project_id=project_id,
        url="",
        title=project.title,
        project=project,
        score=score,
        response_text="AI TZ DRAFT TEXT",
        price="1000",
        delivery_days=7,
        prepared_at=datetime.now(timezone.utc),
        journal_confirmed=True,
        journal_exported=exported,
    )


def test_sync_skips_reappend_when_exported_id_unfindable(tmp_path, monkeypatch) -> None:
    """Confirmed+exported item with empty URL must not append again on second sync."""
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    store = PreparedResponseStore(tmp_path / "prepared")
    item = _prepared_telegram_tz("tz_1785490763952", exported=False)
    store.save(item)

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        lambda settings: (_ for _ in ()).throw(ModuleNotFoundError("mcp")),
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        lambda *a, **k: KworkStatusSyncResult(),
    )
    settings = type("S", (), {"response_journal": str(journal_path)})()

    first = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)
    assert first.appended_prepared == 1
    wb = load_workbook(journal_path)
    rows_after_first = wb.active.max_row
    wb.close()

    second = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)
    assert second.appended_prepared == 0
    wb = load_workbook(journal_path)
    assert wb.active.max_row == rows_after_first
    wb.close()
    saved = store.load("telegram", "tz_manual", "tz_1785490763952")
    assert saved is not None and saved.journal_exported is True


def test_sync_updates_i_from_statedata_comments(tmp_path, monkeypatch) -> None:
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    item = _prepared("3217871", exported=True)
    store = PreparedResponseStore(tmp_path / "prepared")
    store.save(item)
    writer.append_prepared(
        item.project,
        item.score,
        "AI DRAFT TEXT",
        price="1000",
        delivery_days=7,
    )

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        lambda settings: object(),
    )
    monkeypatch.setattr(
        "src.browser.factory.close_browser_client",
        lambda client: None,
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        lambda *a, **k: KworkStatusSyncResult(),
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.fetch_my_offer_comment_details",
        lambda browser, navigate=False: {
            "3217871": KworkOfferComment(
                project_id="3217871",
                comment="REAL KWORK OFFER TEXT " + ("x" * 40),
                price="12000",
                delivery_days=14,
            )
        },
    )

    settings = type("S", (), {"response_journal": str(journal_path)})()
    result = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)

    assert result.updated_notes >= 1
    wb = load_workbook(journal_path)
    cell = str(wb.active.cell(row=2, column=9).value or "")
    assert "REAL KWORK OFFER TEXT" in cell
    assert "AI DRAFT TEXT" not in cell
    assert "12000" in cell


def test_sync_fills_empty_i_from_comments_without_confirm(
    tmp_path, monkeypatch
) -> None:
    """Status-synced row with empty I gets platform text even if prepared is absent."""
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    store = PreparedResponseStore(tmp_path / "prepared")

    writer.append_kwork_offer_status(
        project_id="3233067",
        title="Project 3233067",
        status="Отказ",
        result="Отказ",
        project_type="Telegram-бот",
    )
    assert writer.response_is_empty_by_project_id("3233067") is True

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        lambda settings: object(),
    )
    monkeypatch.setattr(
        "src.browser.factory.close_browser_client",
        lambda client: None,
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        lambda *a, **k: KworkStatusSyncResult(appended=1, matched=1),
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.fetch_my_offer_comment_details",
        lambda browser, navigate=False: {
            "3233067": KworkOfferComment(
                project_id="3233067",
                comment="LIVE OFFER ON KWORK " + ("y" * 40),
                price="5500",
                delivery_days=5,
            )
        },
    )

    settings = type("S", (), {"response_journal": str(journal_path)})()
    result = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)

    assert result.updated_notes >= 1
    assert result.appended_prepared == 0
    expected = format_response_payload(
        "LIVE OFFER ON KWORK " + ("y" * 40),
        price="5500",
        delivery_days=5,
    )
    wb = load_workbook(journal_path)
    cell = str(wb.active.cell(row=2, column=9).value or "")
    wb.close()
    assert cell == expected
    assert "LIVE OFFER ON KWORK" in cell
    assert "5500" in cell
    assert "5 дн." in cell
    assert writer.response_is_empty_by_project_id("3233067") is False


def test_sync_fills_empty_i_when_prepared_unconfirmed(
    tmp_path, monkeypatch
) -> None:
    """Unconfirmed prepared must not block empty-I fill from comments."""
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    store = PreparedResponseStore(tmp_path / "prepared")
    item = _prepared("3233067")
    item.journal_confirmed = False
    store.save(item)

    writer.append_kwork_offer_status(
        project_id="3233067",
        title="Project 3233067",
        status="Отказ",
        result="Отказ",
    )

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        lambda settings: object(),
    )
    monkeypatch.setattr(
        "src.browser.factory.close_browser_client",
        lambda client: None,
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        lambda *a, **k: KworkStatusSyncResult(),
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.fetch_my_offer_comment_details",
        lambda browser, navigate=False: {
            "3233067": KworkOfferComment(
                project_id="3233067",
                comment="PLATFORM TEXT FROM COMMENTS",
                price="8000",
                delivery_days=10,
            )
        },
    )

    settings = type("S", (), {"response_journal": str(journal_path)})()
    result = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)

    assert result.updated_notes >= 1
    assert result.appended_prepared == 0
    wb = load_workbook(journal_path)
    cell = str(wb.active.cell(row=2, column=9).value or "")
    wb.close()
    assert "PLATFORM TEXT FROM COMMENTS" in cell
    assert "8000" in cell
    assert "AI DRAFT TEXT" not in cell
    saved = store.load("kwork", "kwork_dev_it", "3233067")
    assert saved is not None and saved.journal_confirmed is False


def test_empty_fill_path_does_not_overwrite_nonempty_i(
    tmp_path, monkeypatch
) -> None:
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    store = PreparedResponseStore(tmp_path / "prepared")

    writer.append_kwork_offer_status(
        project_id="3233067",
        title="Project 3233067",
        status="Отказ",
        result="Отказ",
    )
    keep = format_response_payload(
        "ALREADY FILLED RESPONSE", price="1000", delivery_days=3
    )
    assert writer.update_response_by_project_id("3233067", keep)

    monkeypatch.setattr(
        "src.browser.factory.get_browser_client",
        lambda settings: object(),
    )
    monkeypatch.setattr(
        "src.browser.factory.close_browser_client",
        lambda client: None,
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.sync_journal_from_kwork_offers",
        lambda *a, **k: KworkStatusSyncResult(),
    )
    monkeypatch.setattr(
        "src.journal.vps_sync.fetch_my_offer_comment_details",
        lambda browser, navigate=False: {
            "3233067": KworkOfferComment(
                project_id="3233067",
                comment="SHOULD NOT REPLACE EXISTING",
                price="99999",
                delivery_days=99,
            )
        },
    )

    settings = type("S", (), {"response_journal": str(journal_path)})()
    result = sync_journal_on_vps(settings=settings, writer=writer, prepared_store=store)

    # Empty-fill path must not touch I; no confirmed prepared either.
    assert result.updated_notes == 0
    wb = load_workbook(journal_path)
    cell = str(wb.active.cell(row=2, column=9).value or "")
    wb.close()
    assert cell == keep
    assert "SHOULD NOT REPLACE EXISTING" not in cell
    assert "99999" not in cell


def test_update_response_by_project_id_if_empty_unit(tmp_path) -> None:
    journal_path = tmp_path / "journal.xlsx"
    writer = JournalWriter(journal_path)
    JournalWriter.create_template_copy(journal_path)
    writer.append_kwork_offer_status(
        project_id="111",
        title="t",
        status="Отказ",
        result="Отказ",
    )
    first = format_response_payload("first", price="1", delivery_days=1)
    assert writer.update_response_by_project_id_if_empty("111", first) is True
    assert writer.update_response_by_project_id_if_empty(
        "111", format_response_payload("second", price="2", delivery_days=2)
    ) is False
    wb = load_workbook(journal_path)
    assert str(wb.active.cell(row=2, column=9).value or "") == first
    wb.close()
