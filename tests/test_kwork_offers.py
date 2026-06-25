from __future__ import annotations

from pathlib import Path

from src.adapters.kwork_offers import (
    journal_status_for_offer,
    parse_offers_html,
    parse_offers_items,
)
from src.journal.kwork_status_sync import sync_journal_from_kwork_offers
from src.journal.writer import JournalWriter
from src.models import GptScoreResult, ProjectFull


def test_parse_offers_items_order_and_waiting() -> None:
    raw = [
        {
            "project_id": "3204427",
            "title": "Доработать 2 Telegram-бота по готовому ТЗ",
            "informers": ["Покупатель сделал 1 заказ"],
        },
        {
            "project_id": "3203948",
            "title": "Парсер",
            "informers": ["Покупатель пока не сделал заказ"],
        },
    ]
    offers = parse_offers_items(raw)
    assert set(offers) == {"3204427", "3203948"}
    assert offers["3204427"].buyer_orders == 1
    assert offers["3203948"].waiting_for_order is True


def test_journal_status_mapping() -> None:
    offers = parse_offers_items(
        [
            {
                "project_id": "1",
                "title": "A",
                "informers": ["Покупатель сделал 2 заказ"],
            },
            {
                "project_id": "2",
                "title": "B",
                "informers": ["Покупатель пока не сделал заказ"],
            },
        ]
    )
    assert journal_status_for_offer(offers["1"]) == (
        "Отказ",
        "Покупатель сделал 2 заказ",
    )
    assert journal_status_for_offer(offers["2"]) == ("Отправлен", "Жду ответа")


def test_parse_offers_html_fixture() -> None:
    html = Path("tests/fixtures/kwork_offers.html").read_text(encoding="utf-8")
    offers = parse_offers_html(html)
    assert offers["3204427"].buyer_orders == 1
    assert offers["3203948"].waiting_for_order is True


def test_sync_journal_from_offers_updates_excel(tmp_path: Path) -> None:
    journal_path = tmp_path / "journal.xlsx"
    JournalWriter.create_template_copy(journal_path)
    project = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3204427",
        url="https://kwork.ru/projects/3204427",
        title="Доработать 2 Telegram-бота по готовому ТЗ",
        full_description="desc",
    )
    score = GptScoreResult(
        score=8,
        fit=True,
        reason="ok",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Telegram-бот",
        competition_level="low",
        recommendation="откликаться",
    )
    writer = JournalWriter(journal_path)
    writer.append_prepared(project, score, "text", price="35000", delivery_days=10)

    offers = parse_offers_items(
        [
            {
                "project_id": "3204427",
                "title": project.title,
                "informers": ["Покупатель сделал 1 заказ"],
            }
        ]
    )
    result = sync_journal_from_kwork_offers(journal_path, offers=offers)
    assert result.error is None
    assert result.matched == 1
    assert result.updated == 1

    from openpyxl import load_workbook

    wb = load_workbook(journal_path)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == "Отказ"
    assert ws.cell(row=2, column=7).value == "Покупатель сделал 1 заказ"
    wb.close()


def test_sync_journal_from_offers_appends_missing_rows(tmp_path: Path) -> None:
    journal_path = tmp_path / "journal.xlsx"
    JournalWriter.create_template_copy(journal_path)
    offers = parse_offers_items(
        [
            {
                "project_id": "3204427",
                "title": "Доработать 2 Telegram-бота по готовому ТЗ",
                "informers": ["Покупатель сделал 1 заказ"],
            },
            {
                "project_id": "3203948",
                "title": "Парсер объявлений по недвижимости",
                "informers": ["Покупатель пока не сделал заказ"],
            },
        ]
    )
    result = sync_journal_from_kwork_offers(journal_path, offers=offers)
    assert result.error is None
    assert result.appended == 2
    assert result.matched == 0

    from openpyxl import load_workbook

    wb = load_workbook(journal_path)
    ws = wb.active
    assert ws.cell(row=2, column=4).value == "https://kwork.ru/projects/3204427/view"
    assert ws.cell(row=2, column=4).hyperlink is not None
    assert ws.cell(row=2, column=5).value == "Telegram-бот"
    assert ws.cell(row=2, column=6).value == "Отказ"
    assert ws.cell(row=3, column=4).value == "https://kwork.ru/projects/3203948/view"
    assert ws.cell(row=3, column=5).value == "Парсинг"
    assert ws.cell(row=3, column=6).value == "Отправлен"
    wb.close()
