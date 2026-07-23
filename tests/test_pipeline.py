from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pytest

from src.config import Settings
from src.models import GptScoreResult, OfferTerms, PendingOffer, ProjectFull, ProjectPreview
from src.pipeline.orchestrator import PipelineOrchestrator
from src.store.repository import ProjectRepository


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    return Settings(
        openai_api_key="test-key",
        openai_base_url="https://api.example.com/openai/v1",
        telegram_bot_token="token",
        telegram_chat_id="1",
        response_journal=str(tmp_path / "journal.xlsx"),
        database_path=str(tmp_path / "test.db"),
        scan_bootstrap_skip_pipeline=False,
        require_telegram_approval=False,
        min_gpt_score=7,
        _env_file=None,
    )


@pytest.fixture
def project_full() -> ProjectFull:
    return ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url="https://kwork.ru/projects/999",
        title="AI Telegram bot for inventory sync",
        full_description="Need a Python/aiogram bot that monitors channels and syncs deals to Sheets.",
        desired_budget="5000",
        offers_count=3,
    )


@pytest.fixture
def score() -> GptScoreResult:
    return GptScoreResult(
        score=9,
        fit=True,
        reason="match",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Telegram-бот",
        competition_level="low",
        recommendation="откликаться",
    )


def _make_orchestrator(
    settings: Settings,
    *,
    previews: list[ProjectPreview],
    project_full: ProjectFull,
    score: GptScoreResult,
) -> PipelineOrchestrator:
    repo = ProjectRepository(settings.database_path)

    mock_adapter = MagicMock()
    mock_adapter.scan_new.return_value = previews
    mock_adapter.read_full.return_value = project_full
    mock_adapter.submit_response.return_value = MagicMock(
        success=True, project_id="999", message="ok"
    )
    mock_adapter.prepare_response.return_value = MagicMock(
        success=True, project_id="999", message="prepared"
    )

    mock_estimator = MagicMock()
    mock_estimator.estimate.return_value = OfferTerms(
        price_rub=5000, delivery_days=14, plan_summary=""
    )
    mock_estimator.estimate_market_cost.return_value = 5000
    mock_estimator.close = MagicMock()

    mock_scorer = MagicMock()
    mock_scorer.score.return_value = score
    mock_scorer.close = MagicMock()

    mock_generator = MagicMock()
    mock_generator.generate.return_value = "Generated response text"
    mock_generator.close = MagicMock()

    mock_lightrag = MagicMock()
    mock_lightrag.get_full_context.return_value = "ctx"

    mock_tg = MagicMock()
    mock_tg.send_review_card = AsyncMock(return_value=1)
    mock_tg.send_offer_link = AsyncMock(return_value=2)
    mock_tg.send_form_prepared_ready = AsyncMock(return_value=3)
    mock_tg.mark_review_skipped = AsyncMock()
    mock_tg.mark_review_approved = AsyncMock()
    mock_tg.send_photo = AsyncMock()
    mock_tg.notify = AsyncMock()
    mock_tg.close = AsyncMock()

    from src.journal.writer import JournalWriter
    from src.responses.prepared_store import PreparedResponseStore
    from src.telegram_bot.pending_store import PendingStore
    from src.telegram_bot.review_service import ReviewService

    store = PendingStore(base_dir=Path(settings.database_path).parent / "pending")
    review = ReviewService(settings, store, mock_tg, repo)

    orch = PipelineOrchestrator(
        settings=settings,
        repository=repo,
        review_service=review,
        scorer=mock_scorer,
        response_generator=mock_generator,
        lightrag=mock_lightrag,
        journal=JournalWriter(settings.response_journal),
        prepared_store=PreparedResponseStore(
            Path(settings.database_path).parent / "prepared"
        ),
        offer_estimator=mock_estimator,
        adapter_factory=lambda _s, _b=None: mock_adapter,
        browser=MagicMock(),
    )
    return orch, mock_adapter


@pytest.mark.asyncio
async def test_pipeline_scan_score_submit(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
    )
    orch, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )
    totals = await orch.run_scan_cycle()

    assert totals["new"] == 1
    orch.scorer.score.assert_called_once()
    orch.response_generator.generate.assert_called_once()
    mock_adapter.prepare_response.assert_called_once()

    repo = ProjectRepository(settings.database_path)
    assert repo.is_known("kwork", "kwork_dev_it", "999")
    state = repo.get_scan_state("kwork_dev_it")
    assert state is not None


def test_price_exceeds_budget_ceiling() -> None:
    from src.adapters.kwork_pricing import price_exceeds_budget_ceiling
    from src.models import ProjectFull

    project = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="1",
        url="https://kwork.ru/projects/1/view",
        title="Test",
        full_description="desc",
        max_budget="до 25 000 ₽",
    )
    assert price_exceeds_budget_ceiling(60_000, project, multiplier=2.0) is True
    assert price_exceeds_budget_ceiling(50_000, project, multiplier=2.0) is False
    assert price_exceeds_budget_ceiling(50_001, project, multiplier=2.0) is True

    no_ceiling = project.model_copy(update={"max_budget": None})
    assert price_exceeds_budget_ceiling(100_000, no_ceiling) is False


@pytest.mark.asyncio
async def test_pipeline_skips_over_budget_ceiling(
    settings: Settings, score: GptScoreResult
) -> None:
    project_full = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="1001",
        url="https://kwork.ru/projects/1001/view",
        title="Heavy project",
        full_description="Big integration",
        max_budget="до 25 000 ₽",
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="1001",
        url=project_full.url,
        title=project_full.title,
    )
    orch, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )
    orch.offer_estimator.estimate_market_cost.return_value = 60_000

    totals = await orch.run_scan_cycle()

    assert totals["new"] == 1
    orch.review_service.tg_bot.send_review_card.assert_not_called()
    mock_adapter.prepare_response.assert_not_called()
    repo = ProjectRepository(settings.database_path)
    with repo._conn() as conn:
        row = conn.execute(
            "SELECT status FROM projects WHERE project_id = ?",
            ("1001",),
        ).fetchone()
    assert row is not None
    assert row["status"] == "skipped"


@pytest.mark.asyncio
async def test_pipeline_extract_fail_skips_scoring(
    settings: Settings, score: GptScoreResult
) -> None:
    project_full = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="2002",
        url="https://kwork.ru/projects/2002/view",
        title="",
        full_description="",
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="2002",
        url=project_full.url,
        # Wrong listing title must not resurrect failed page extract
        title="Отметки пользователей в Telegram Stories",
    )
    orch, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )
    totals = await orch.run_scan_cycle()

    assert totals["new"] == 1
    orch.scorer.score.assert_not_called()
    orch.review_service.tg_bot.send_review_card.assert_not_called()
    mock_adapter.prepare_response.assert_not_called()
    repo = ProjectRepository(settings.database_path)
    assert repo.is_known("kwork", "kwork_dev_it", "2002")


@pytest.mark.asyncio
async def test_pipeline_extract_fail_weak_title_as_desc(
    settings: Settings, score: GptScoreResult
) -> None:
    title = "Нужен Telegram-бот для учёта заявок и CRM интеграции"
    assert len(title) > 40
    project_full = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3217391",
        url="https://kwork.ru/projects/3217391/view",
        title=title,
        full_description=title,
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3217391",
        url=project_full.url,
        title=title,
    )
    orch, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )
    totals = await orch.run_scan_cycle()

    assert totals["new"] == 1
    orch.scorer.score.assert_not_called()
    orch.review_service.tg_bot.send_review_card.assert_not_called()
    mock_adapter.prepare_response.assert_not_called()


@pytest.mark.asyncio
async def test_pipeline_title_mismatch_keeps_page_and_scores(
    settings: Settings, score: GptScoreResult
) -> None:
    project_full = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="2003",
        url="https://kwork.ru/projects/2003/view",
        title="Ретушь фото для каталога",
        full_description="Нужна цветокоррекция и подготовка 40 фото для маркетплейса.",
        desired_budget="до 3 000 ₽",
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="2003",
        url=project_full.url,
        title="Отметки пользователей в Telegram Stories",
    )
    orch, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )
    totals = await orch.run_scan_cycle()

    assert totals["new"] == 1
    orch.scorer.score.assert_called_once()
    scored = orch.scorer.score.call_args[0][0]
    assert scored.title == "Ретушь фото для каталога"


@pytest.mark.asyncio
async def test_pipeline_quick_win_skips_budget_ceiling(
    settings: Settings,
) -> None:
    quick_project = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3211980",
        url="https://kwork.ru/projects/3211980/view",
        title="Парсинг сайта",
        full_description="Небольшой парсинг, выгрузка в csv",
        max_budget="до 3 000 ₽",
        offers_count=5,
    )
    quick_score = GptScoreResult(
        score=6,
        fit=True,
        reason="quick script",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Парсинг",
        competition_level="low",
        recommendation="откликаться",
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3211980",
        url=quick_project.url,
        title=quick_project.title,
    )
    settings.require_telegram_approval = True
    orch, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=quick_project, score=quick_score
    )
    orch.offer_estimator.estimate_market_cost.return_value = 60_000

    totals = await orch.run_scan_cycle()

    assert totals["new"] == 1
    orch.review_service.tg_bot.send_review_card.assert_called_once()
    mock_adapter.prepare_response.assert_not_called()


@pytest.mark.asyncio
async def test_pipeline_skips_low_score(
    settings: Settings, project_full: ProjectFull
) -> None:
    low_score = GptScoreResult(
        score=4,
        fit=False,
        reason="no",
        matched_skills=[],
        risks=[],
        suggested_project_type="X",
        competition_level="high",
        recommendation="пропустить",
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="888",
        url="https://kwork.ru/projects/888",
        title="Low",
    )
    orch, _ = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=low_score
    )
    await orch.run_scan_cycle()
    orch.response_generator.generate.assert_not_called()


@pytest.mark.asyncio
async def test_handle_approved_requires_approval_flag(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.require_telegram_approval = True
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="777",
        url=project_full.url,
        title="x",
    )
    orch, _ = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )

    pending = PendingOffer(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="777",
        url=project_full.url,
        title=project_full.title,
        project=project_full,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="pending",
    )
    await orch.handle_approved("kwork", "kwork_dev_it", "777", pending)
    orch.response_generator.generate.assert_not_called()


@pytest.mark.asyncio
async def test_handle_approved_unknown_source_notifies(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.require_telegram_approval = False
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="666",
        url=project_full.url,
        title="x",
    )
    orch, _ = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )

    offer = PendingOffer(
        platform="kwork",
        source_key="unknown_source",
        project_id="666",
        url=project_full.url,
        title=project_full.title,
        project=project_full,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text="text",
    )
    await orch.handle_approved("kwork", "unknown_source", "666", offer)

    orch.review_service.tg_bot.notify.assert_called_once()
    assert "unknown_source" in orch.review_service.tg_bot.notify.call_args[0][0]


@pytest.mark.asyncio
async def test_journal_confirm_writes_excel(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    from src.adapters.kwork import OfferFormSnapshot
    from src.responses.prepared_store import PreparedResponse, PreparedResponseStore
    from unittest.mock import MagicMock

    orch, _ = _make_orchestrator(
        settings, previews=[], project_full=project_full, score=score
    )
    store = PreparedResponseStore(
        Path(settings.database_path).parent / "prepared"
    )
    orch.prepared_store = store
    item = PreparedResponse(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
        project=project_full,
        score=score,
        response_text="text",
        price="5000",
        delivery_days=7,
    )
    store.save(item)

    callback = MagicMock()
    callback.answer = AsyncMock()
    orch.review_service.tg_bot.mark_journal_confirmed = AsyncMock()
    orch.review_service.tg_bot.notify = AsyncMock()
    orch._fetch_submitted_offer_text = MagicMock(
        return_value=OfferFormSnapshot(description="", ok=False, error="skip")
    )

    await orch.handle_journal_confirm(
        "kwork", "kwork_dev_it", "999", callback
    )

    saved = store.load("kwork", "kwork_dev_it", "999")
    assert saved is not None
    assert saved.journal_confirmed is True
    assert saved.journal_exported is True
    from openpyxl import load_workbook

    wb = load_workbook(settings.response_journal)
    assert wb.active.max_row >= 2


@pytest.mark.asyncio
async def test_journal_confirm_uses_kwork_platform_text(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    from src.adapters.kwork import OfferFormSnapshot
    from src.responses.prepared_store import PreparedResponse, PreparedResponseStore
    from openpyxl import load_workbook
    from unittest.mock import MagicMock

    orch, _ = _make_orchestrator(
        settings, previews=[], project_full=project_full, score=score
    )
    store = PreparedResponseStore(
        Path(settings.database_path).parent / "prepared"
    )
    orch.prepared_store = store
    item = PreparedResponse(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
        project=project_full,
        score=score,
        response_text="AI ONLY DRAFT",
        price="5000",
        delivery_days=7,
    )
    store.save(item)

    callback = MagicMock()
    callback.answer = AsyncMock()
    orch.review_service.tg_bot.mark_journal_confirmed = AsyncMock()
    orch.review_service.tg_bot.notify = AsyncMock()
    platform_text = "Edited offer text from Kwork form " + ("y" * 40)
    orch._fetch_submitted_offer_text = MagicMock(
        return_value=OfferFormSnapshot(
            description=platform_text,
            price="8800",
            delivery_days=10,
            ok=True,
        )
    )

    await orch.handle_journal_confirm(
        "kwork", "kwork_dev_it", "999", callback
    )

    saved = store.load("kwork", "kwork_dev_it", "999")
    assert saved is not None
    assert saved.response_text == platform_text
    assert saved.price == "8800"
    assert saved.delivery_days == 10

    wb = load_workbook(settings.response_journal)
    cell = str(wb.active.cell(row=2, column=9).value or "")
    assert "Edited offer text from Kwork form" in cell
    assert "AI ONLY DRAFT" not in cell
    assert "8800" in cell


@pytest.mark.asyncio
async def test_journal_confirm_skips_kwork_snapshot_for_non_kwork(
    settings: Settings, score: GptScoreResult
) -> None:
    from src.responses.prepared_store import PreparedResponse, PreparedResponseStore
    from unittest.mock import MagicMock

    yandex_id = "342870b9-bea4-40b7-a616-81061dc845db"
    project = ProjectFull(
        platform="yandex_uslugi",
        source_key="yandex_uslugi_it",
        project_id=yandex_id,
        url=f"https://uslugi.yandex.ru/order/{yandex_id}",
        title="Yandex order",
        full_description="desc",
    )
    orch, _ = _make_orchestrator(
        settings, previews=[], project_full=project, score=score
    )
    store = PreparedResponseStore(
        Path(settings.database_path).parent / "prepared"
    )
    orch.prepared_store = store
    item = PreparedResponse(
        platform="yandex_uslugi",
        source_key="yandex_uslugi_it",
        project_id=yandex_id,
        url=project.url,
        title=project.title,
        project=project,
        score=score,
        response_text="Yandex prepared text",
        price="3000",
        delivery_days=5,
    )
    store.save(item)

    callback = MagicMock()
    callback.answer = AsyncMock()
    orch.review_service.tg_bot.mark_journal_confirmed = AsyncMock()
    orch.review_service.tg_bot.notify = AsyncMock()
    orch._fetch_submitted_offer_text = MagicMock(
        side_effect=AssertionError("kwork snapshot must not run for yandex")
    )

    await orch.handle_journal_confirm(
        "yandex_uslugi", "yandex_uslugi_it", yandex_id, callback
    )

    orch._fetch_submitted_offer_text.assert_not_called()
    saved = store.load("yandex_uslugi", "yandex_uslugi_it", yandex_id)
    assert saved is not None
    assert saved.journal_exported is True
    assert saved.response_text == "Yandex prepared text"

    from openpyxl import load_workbook

    wb = load_workbook(settings.response_journal)
    assert wb.active.cell(row=2, column=6).value == "Отправлен"
    assert wb.active.cell(row=2, column=7).value == "Жду ответа"
    wb.close()


@pytest.mark.asyncio
async def test_prepare_success_always_notifies(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.prepare_only_no_submit = True
    orch, mock_adapter = _make_orchestrator(
        settings,
        previews=[],
        project_full=project_full,
        score=score,
    )
    mock_adapter.prepare_response.return_value = MagicMock(
        success=True,
        project_id=project_full.project_id,
        message="prepared: form filled, submit not clicked",
    )
    offer = PendingOffer(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
        project=project_full,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text="Готовый текст отклика для формы.",
    )
    await orch._prepare_offer_on_site(offer)
    orch.review_service.tg_bot.send_form_prepared_ready.assert_called_once()
    call_kw = orch.review_service.tg_bot.send_form_prepared_ready.call_args.kwargs
    assert "new_offer?project=" in call_kw["offer_url"]


@pytest.mark.asyncio
async def test_process_manual_kwork_project_sends_review_card(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.require_telegram_approval = True
    orch, _ = _make_orchestrator(
        settings,
        previews=[],
        project_full=project_full,
        score=score,
    )

    result = await orch.process_manual_kwork_project("123456")

    assert result == {"project_id": "123456", "outcome": "notified"}
    orch.review_service.tg_bot.send_review_card.assert_called_once()
    offer = orch.review_service.tg_bot.send_review_card.call_args.args[0]
    assert offer.source_key == "kwork_manual"
    assert orch.repository.is_known("kwork", "kwork_manual", "123456")


@pytest.mark.asyncio
async def test_prepare_offer_resolves_kwork_manual_source(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.prepare_only_no_submit = True
    orch, mock_adapter = _make_orchestrator(
        settings,
        previews=[],
        project_full=project_full,
        score=score,
    )
    mock_adapter.prepare_response.return_value = MagicMock(
        success=True,
        project_id=project_full.project_id,
        message="prepared",
    )
    manual_project = project_full.model_copy(update={"source_key": "kwork_manual"})
    offer = PendingOffer(
        platform="kwork",
        source_key="kwork_manual",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
        project=manual_project,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text="Текст отклика.",
    )
    await orch._prepare_offer_on_site(offer)
    mock_adapter.prepare_response.assert_called_once()


@pytest.mark.asyncio
async def test_prepare_uses_same_commercial_price_for_tg_and_form(
    settings: Settings, score: GptScoreResult
) -> None:
    """Market 40k + offer terms 60k → both TG Оценка and form use min (40k)."""
    settings.prepare_only_no_submit = True
    project = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="dual_price",
        url="https://kwork.ru/projects/dual_price",
        title="Dual price project",
        full_description="Python bot with integrations.",
        desired_budget="до 20 000 ₽",
        max_budget="до 100 000 ₽",
    )
    orch, mock_adapter = _make_orchestrator(
        settings,
        previews=[],
        project_full=project,
        score=score,
    )
    orch.offer_estimator.estimate.return_value = OfferTerms(
        price_rub=60_000, delivery_days=10, plan_summary=""
    )
    orch.offer_estimator.estimate_market_cost.return_value = 40_000
    mock_adapter.prepare_response.return_value = MagicMock(
        success=True,
        project_id="dual_price",
        message="prepared",
    )
    offer = PendingOffer(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="dual_price",
        url=project.url,
        title=project.title,
        project=project,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text="Срок — 10 дней. Стоимость — от 40 000 ₽.",
    )
    await orch._prepare_offer_on_site(offer)
    prepared_price = mock_adapter.prepare_response.call_args.args[2]
    assert prepared_price == "40000"
    ready_kw = orch.review_service.tg_bot.send_form_prepared_ready.call_args.kwargs
    assert ready_kw["price"] == "40000"
    notify_text = " ".join(
        str(c.args[0]) for c in orch.review_service.tg_bot.notify.await_args_list
    )
    assert "40 000" in notify_text
    assert "60 000" not in notify_text


@pytest.mark.asyncio
async def test_prepare_budget_gap_clamps_form_price_and_appends_note(
    settings: Settings, score: GptScoreResult
) -> None:
    settings.prepare_only_no_submit = True
    cheap = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url="https://kwork.ru/projects/999",
        title="Big bot tiny budget",
        full_description="Нужен сложный AI Telegram-бот с интеграциями.",
        max_budget="до 1 500 ₽",
    )
    orch, mock_adapter = _make_orchestrator(
        settings,
        previews=[],
        project_full=cheap,
        score=score,
    )
    orch.offer_estimator.estimate.return_value = OfferTerms(
        price_rub=1500, delivery_days=14, plan_summary=""
    )
    orch.offer_estimator.estimate_market_cost.return_value = 20_000
    mock_adapter.prepare_response.return_value = MagicMock(
        success=True,
        project_id="999",
        message="prepared",
    )
    offer = PendingOffer(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=cheap.url,
        title=cheap.title,
        project=cheap,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text=(
            "Соберу бота под задачу.\n"
            "Срок — 14 дней. Стоимость — от 20 000 ₽.\n"
            "Предлагаю обсудить детали и приступить."
        ),
    )
    await orch._prepare_offer_on_site(offer)
    mock_adapter.prepare_response.assert_called_once()
    args = mock_adapter.prepare_response.call_args
    # prepare_response(project_id, text, price, ...)
    prepared_text = args.args[1] if args.args else args.kwargs.get("text")
    prepared_price = args.args[2] if len(args.args) > 2 else args.kwargs.get("price")
    assert prepared_price == "1500"
    assert "обсудить сумму" in prepared_text.lower()
    assert "занижен" in prepared_text.lower()
    notify_text = " ".join(
        str(c.args[0]) for c in orch.review_service.tg_bot.notify.await_args_list
    )
    assert "20 000" in notify_text or "20000" in notify_text.replace(" ", "")
    assert "1500" in notify_text.replace(" ", "") or "1 500" in notify_text


@pytest.mark.asyncio
async def test_prepare_budget_gap_from_desired_only_shows_ceiling_in_tg(
    settings: Settings, score: GptScoreResult
) -> None:
    """3218308: max_budget None, desired «до 1 500» — still gap + потолок in TG."""
    settings.prepare_only_no_submit = True
    cheap = ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3218308",
        url="https://kwork.ru/projects/3218308",
        title="Checko Python",
        full_description="Интеграция Checko API на Python.",
        desired_budget="до 1 500 ₽",
        max_budget=None,
    )
    orch, mock_adapter = _make_orchestrator(
        settings,
        previews=[],
        project_full=cheap,
        score=score,
    )
    orch.offer_estimator.estimate.return_value = OfferTerms(
        price_rub=25_000, delivery_days=7, plan_summary=""
    )
    orch.offer_estimator.estimate_market_cost.return_value = 25_000
    mock_adapter.prepare_response.return_value = MagicMock(
        success=True,
        project_id="3218308",
        message="prepared",
    )
    offer = PendingOffer(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="3218308",
        url=cheap.url,
        title=cheap.title,
        project=cheap,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text="Срок — 7 дней. Стоимость — от 25 000 ₽.",
    )
    await orch._prepare_offer_on_site(offer)
    args = mock_adapter.prepare_response.call_args
    prepared_text = args.args[1] if args.args else args.kwargs.get("text")
    prepared_price = args.args[2] if len(args.args) > 2 else args.kwargs.get("price")
    assert prepared_price == "1500"
    assert "обсудить сумму" in prepared_text.lower()
    notify_text = " ".join(
        str(c.args[0]) for c in orch.review_service.tg_bot.notify.await_args_list
    )
    assert "потолок" in notify_text.lower()
    assert "1 500" in notify_text or "1500" in notify_text.replace(" ", "")


@pytest.mark.asyncio
async def test_process_manual_kwork_project_skips_score_gate(
    settings: Settings, project_full: ProjectFull
) -> None:
    low_score = GptScoreResult(
        score=2,
        fit=False,
        reason="no match",
        matched_skills=[],
        risks=["out of stack"],
        suggested_project_type="X",
        competition_level="high",
        recommendation="пропустить",
    )
    settings.require_telegram_approval = True
    orch, _ = _make_orchestrator(
        settings,
        previews=[],
        project_full=project_full,
        score=low_score,
    )

    result = await orch.process_manual_kwork_project("123456")

    assert result["outcome"] == "notified"
    orch.review_service.tg_bot.send_review_card.assert_called_once()
