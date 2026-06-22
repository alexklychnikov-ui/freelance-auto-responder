from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from src.config import Settings
from src.journal.writer import JOURNAL_COLUMNS, JournalWriter
from src.models import GptScoreResult, PendingOffer, ProjectFull, ProjectPreview
from src.pipeline.orchestrator import PipelineOrchestrator
from src.store.repository import ProjectRepository


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    return Settings(
        openai_api_key="test-key",
        openai_base_url="https://api.example.com/openai/v1",
        telegram_bot_token="token",
        telegram_chat_id="1",
        response_journal=str(tmp_path / "journal.xlsx"),
        database_path=str(tmp_path / "test.db"),
        scan_bootstrap_skip_pipeline=True,
        require_telegram_approval=True,
        min_gpt_score=7,
        max_daily_responses=5,
        _env_file=None,
    )


@pytest.fixture
def project_full() -> ProjectFull:
    return ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url="https://kwork.ru/projects/999",
        title="AI bot",
        full_description="Need Python bot",
        desired_budget="5000",
        offers_count=3,
    )


@pytest.fixture
def score() -> GptScoreResult:
    return GptScoreResult(
        score=9,
        fit=True,
        reason="match",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Telegram-бот",
        competition_level="low",
        recommendation="откликаться",
    )


def _project_status(
    repo: ProjectRepository, platform: str, source_key: str, project_id: str
) -> str | None:
    with repo._conn() as conn:
        row = conn.execute(
            """
            SELECT status FROM projects
            WHERE platform = ? AND source_key = ? AND project_id = ?
            """,
            (platform, source_key, project_id),
        ).fetchone()
    return row["status"] if row else None


def _make_orchestrator(
    settings: Settings,
    *,
    previews: list[ProjectPreview],
    project_full: ProjectFull,
    score: GptScoreResult,
) -> tuple[PipelineOrchestrator, MagicMock, MagicMock, MagicMock]:
    repo = ProjectRepository(settings.database_path)

    mock_adapter = MagicMock()
    mock_adapter.scan_new.return_value = previews
    mock_adapter.read_full.return_value = project_full
    mock_adapter.submit_response.return_value = MagicMock(
        success=True, project_id=project_full.project_id, message="ok"
    )

    mock_scorer = MagicMock()
    mock_scorer.score.return_value = score
    mock_scorer.close = MagicMock()

    mock_generator = MagicMock()
    mock_generator.generate.return_value = "Generated response text"
    mock_generator.close = MagicMock()

    mock_lightrag = MagicMock()
    mock_lightrag.get_full_context.return_value = "ctx"

    mock_tg = MagicMock()
    mock_tg.send_review_card = AsyncMock(return_value=1)
    mock_tg.send_draft_for_edit = AsyncMock(return_value=2)
    mock_tg.mark_review_skipped = AsyncMock()
    mock_tg.mark_review_approved = AsyncMock()
    mock_tg.send_photo = AsyncMock()
    mock_tg.notify = AsyncMock()
    mock_tg.close = AsyncMock()

    from src.journal.writer import JournalWriter
    from src.responses.prepared_store import PreparedResponseStore
    from src.telegram_bot.pending_store import PendingStore
    from src.telegram_bot.review_service import ReviewService

    store = PendingStore(base_dir=Path(settings.database_path).parent / "pending")
    review = ReviewService(settings, store, mock_tg, repo)

    orch = PipelineOrchestrator(
        settings=settings,
        repository=repo,
        review_service=review,
        scorer=mock_scorer,
        response_generator=mock_generator,
        lightrag=mock_lightrag,
        journal=JournalWriter(settings.response_journal),
        prepared_store=PreparedResponseStore(
            Path(settings.database_path).parent / "prepared"
        ),
        adapter_factory=lambda _s, _b=None: mock_adapter,
        browser=MagicMock(),
    )
    return orch, mock_scorer, mock_tg, mock_adapter


def _journal_with_today_rows(path: Path, count: int) -> None:
    JournalWriter.create_template_copy(path)
    wb = load_workbook(path)
    ws = wb.active
    today = date.today().isoformat()
    for i in range(count):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=today)
        for col, header in enumerate(JOURNAL_COLUMNS[2:], start=3):
            ws.cell(row=row, column=col, value=header)
    wb.save(path)
    wb.close()


@pytest.mark.asyncio
async def test_bootstrap_skip_no_gpt_or_tg_calls(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    previews = [
        ProjectPreview(
            platform="kwork",
            source_key="kwork_dev_it",
            project_id="101",
            url="https://kwork.ru/projects/101",
            title="Bootstrap 1",
        ),
        ProjectPreview(
            platform="kwork",
            source_key="kwork_dev_it",
            project_id="102",
            url="https://kwork.ru/projects/102",
            title="Bootstrap 2",
        ),
    ]
    orch, mock_scorer, mock_tg, _ = _make_orchestrator(
        settings, previews=previews, project_full=project_full, score=score
    )

    totals = await orch.run_scan_cycle()

    assert totals["new"] == 0
    assert totals["skipped"] == 2
    mock_scorer.score.assert_not_called()
    mock_tg.send_review_card.assert_not_called()
    mock_tg.notify.assert_not_called()

    repo = ProjectRepository(settings.database_path)
    assert _project_status(repo, "kwork", "kwork_dev_it", "101") == "skipped"
    assert _project_status(repo, "kwork", "kwork_dev_it", "102") == "skipped"


@pytest.mark.asyncio
async def test_known_project_rescan_skipped(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.scan_bootstrap_skip_pipeline = False
    repo = ProjectRepository(settings.database_path)
    repo.insert_new(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        title=project_full.title,
        url=project_full.url,
        status="scored",
    )
    repo.set_scan_state(
        source_key="kwork_dev_it",
        platform="kwork",
        last_known_project_id="999",
    )

    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
    )
    orch, mock_scorer, mock_tg, mock_adapter = _make_orchestrator(
        settings, previews=[preview], project_full=project_full, score=score
    )

    totals = await orch.run_scan_cycle()

    assert totals["skipped"] >= 1
    mock_scorer.score.assert_not_called()
    mock_adapter.read_full.assert_not_called()
    mock_tg.send_review_card.assert_not_called()


@pytest.mark.asyncio
async def test_gpt_reject_no_send_review_card(
    settings: Settings, project_full: ProjectFull
) -> None:
    settings.scan_bootstrap_skip_pipeline = False
    low_score = GptScoreResult(
        score=4,
        fit=False,
        reason="no",
        matched_skills=[],
        risks=[],
        suggested_project_type="X",
        competition_level="high",
        recommendation="пропустить",
    )
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="888",
        url="https://kwork.ru/projects/888",
        title="Low",
    )
    project_low = project_full.model_copy(
        update={
            "project_id": "888",
            "url": "https://kwork.ru/projects/888",
            "title": "Low",
        }
    )
    orch, _, mock_tg, _ = _make_orchestrator(
        settings, previews=[preview], project_full=project_low, score=low_score
    )

    await orch.run_scan_cycle()

    mock_tg.send_review_card.assert_not_called()
    repo = ProjectRepository(settings.database_path)
    assert _project_status(repo, "kwork", "kwork_dev_it", "888") == "skipped"


@pytest.mark.asyncio
async def test_daily_limit_blocks_submit(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult, tmp_path: Path
) -> None:
    journal = tmp_path / "full_journal.xlsx"
    _journal_with_today_rows(journal, 5)
    settings.response_journal = str(journal)
    settings.require_telegram_approval = False

    orch, _, mock_tg, mock_adapter = _make_orchestrator(
        settings, previews=[], project_full=project_full, score=score
    )

    offer = PendingOffer(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="999",
        url=project_full.url,
        title=project_full.title,
        project=project_full,
        score=score,
        created_at=datetime.now(timezone.utc),
        status="approved",
        approved_at=datetime.now(timezone.utc),
        response_text="Ready response",
    )
    await orch.handle_approved("kwork", "kwork_dev_it", "999", offer)

    mock_adapter.submit_response.assert_not_called()
    mock_tg.notify.assert_called_once()
    assert "лимит" in mock_tg.notify.call_args[0][0].lower()


@pytest.mark.asyncio
async def test_reject_then_rescan_no_second_notify(
    settings: Settings, project_full: ProjectFull, score: GptScoreResult
) -> None:
    settings.scan_bootstrap_skip_pipeline = False
    preview = ProjectPreview(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="555",
        url="https://kwork.ru/projects/555",
        title="Review me",
    )
    project_review = project_full.model_copy(
        update={
            "project_id": "555",
            "url": "https://kwork.ru/projects/555",
            "title": "Review me",
        }
    )
    orch, _, mock_tg, _ = _make_orchestrator(
        settings, previews=[preview], project_full=project_review, score=score
    )
    repo = ProjectRepository(settings.database_path)
    repo.set_scan_state(
        source_key="kwork_dev_it",
        platform="kwork",
        last_known_project_id="554",
    )

    await orch.run_scan_cycle()
    assert mock_tg.send_review_card.call_count == 1

    callback = MagicMock()
    callback.answer = AsyncMock()
    await orch.review_service._handle_reject(
        "kwork", "kwork_dev_it", "555", callback
    )
    assert _project_status(repo, "kwork", "kwork_dev_it", "555") == "rejected"

    await orch.run_scan_cycle()
    assert mock_tg.send_review_card.call_count == 1
