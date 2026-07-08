from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.journal.writer import JOURNAL_COLUMNS, JournalWriter
from src.models import GptScoreResult, ProjectFull


@pytest.fixture
def journal_path(tmp_path: Path) -> Path:
    path = tmp_path / "journal.xlsx"
    JournalWriter.create_template_copy(path)
    return path


@pytest.fixture
def project() -> ProjectFull:
    return ProjectFull(
        platform="kwork",
        source_key="kwork_dev_it",
        project_id="100",
        url="https://kwork.ru/projects/100",
        title="Test project",
        full_description="desc",
    )


@pytest.fixture
def score() -> GptScoreResult:
    return GptScoreResult(
        score=8,
        fit=True,
        reason="ok",
        matched_skills=["Python"],
        risks=[],
        suggested_project_type="Telegram-бот",
        competition_level="low",
        recommendation="откликаться",
    )


def test_journal_template_has_columns(journal_path: Path) -> None:
    wb = load_workbook(journal_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(JOURNAL_COLUMNS) + 1)]
    assert headers == JOURNAL_COLUMNS
    wb.close()


def test_journal_append_row(journal_path: Path, project: ProjectFull, score: GptScoreResult) -> None:
    writer = JournalWriter(journal_path)
    row = writer.append_submission(project, score, "Текст отклика для теста")

    wb = load_workbook(journal_path)
    ws = wb.active
    assert row == 2
    cell_date = ws.cell(row=2, column=2).value
    expected = date.today()
    if isinstance(cell_date, datetime):
        assert cell_date.date() == expected
    else:
        assert cell_date == expected.isoformat()
    assert ws.cell(row=2, column=3).value == "Kwork"
    assert ws.cell(row=2, column=4).value == project.url
    assert ws.cell(row=2, column=4).hyperlink is not None
    assert ws.cell(row=2, column=5).value == "Telegram-бот"
    assert ws.cell(row=2, column=6).value == "Отправлен"
    wb.close()


def test_journal_skips_merged_cells(
    journal_path: Path, project: ProjectFull, score: GptScoreResult
) -> None:
    wb = load_workbook(journal_path)
    ws = wb.active
    ws.merge_cells("A2:I2")
    wb.save(journal_path)
    wb.close()

    writer = JournalWriter(journal_path)
    row = writer.append_prepared(project, score, "prepared text", price="5000", delivery_days=14)

    assert row == 3
    wb = load_workbook(journal_path)
    ws = wb.active
    assert ws.cell(row=3, column=6).value == "Подготовлен"
    assert ws.cell(row=3, column=8).value == "desc"
    assert "prepared text" in str(ws.cell(row=3, column=9).value)
    assert "5000" in str(ws.cell(row=3, column=9).value)
    wb.close()


def test_infer_project_type() -> None:
    from src.journal.writer import infer_project_type

    assert infer_project_type("Доработать 2 Telegram-бота") == "Telegram-бот"
    assert infer_project_type("Парсер объявлений") == "Парсинг"
    assert infer_project_type("Сайт на WordPress") == "Веб-MVP"
    assert infer_project_type("Что-то непонятное") == "Другое"


def test_repair_row_fills_type_and_hyperlink(tmp_path: Path) -> None:
    journal_path = tmp_path / "journal.xlsx"
    JournalWriter.create_template_copy(journal_path)
    wb = load_workbook(journal_path)
    ws = wb.active
    ws.cell(row=2, column=4, value="https://kwork.ru/projects/3204427")
    ws.cell(row=2, column=5, value="—")
    ws.cell(row=2, column=8, value="Доработать 2 Telegram-бота по готовому ТЗ")
    wb.save(journal_path)
    wb.close()

    writer = JournalWriter(journal_path)
    assert writer.repair_row_by_project_id("3204427") is True

    wb = load_workbook(journal_path)
    ws = wb.active
    assert ws.cell(row=2, column=4).hyperlink is not None
    assert ws.cell(row=2, column=5).value == "Telegram-бот"
    wb.close()


def test_format_response_payload() -> None:
    from src.journal.writer import format_response_payload

    text = format_response_payload("Сделаю за вечер", price="8000", delivery_days=10)
    assert text == "Сделаю за вечер\n\nОбщий бюджет: 8000 ₽\nСрок: 10 дн."


def test_project_ids_in_journal(journal_path: Path, project: ProjectFull, score: GptScoreResult) -> None:
    writer = JournalWriter(journal_path)
    assert writer.project_ids_in_journal() == set()

    writer.append_prepared(project, score, "text", price="1000")
    assert writer.project_ids_in_journal() == {"100"}


def test_update_status_by_project_id(
    journal_path: Path, project: ProjectFull, score: GptScoreResult
) -> None:
    writer = JournalWriter(journal_path)
    writer.append_prepared(project, score, "text", price="1000")

    changed = writer.update_status_by_project_id(
        project.project_id,
        status="Отправлен",
        result="Жду ответа",
    )
    assert changed is True

    wb = load_workbook(journal_path)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == "Отправлен"
    assert ws.cell(row=2, column=7).value == "Жду ответа"
    wb.close()

    assert (
        writer.update_status_by_project_id(
            project.project_id,
            status="Отправлен",
            result="Жду ответа",
        )
        is False
    )


def test_journal_template_header_row(
    tmp_path: Path, project: ProjectFull, score: GptScoreResult
) -> None:
    from openpyxl import Workbook

    path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=4, column=2, value="Дата отклика")
    ws.cell(row=4, column=1, value="№")
    ws.cell(row=5, column=1, value=1)
    ws.cell(row=5, column=2, value=date.today())
    ws.cell(row=5, column=4, value="https://example.com/1")
    wb.save(path)
    wb.close()

    writer = JournalWriter(path)
    row = writer.append_prepared(project, score, "text", price="1000")

    assert row == 6
    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(row=6, column=1).value == 2
    assert ws.cell(row=6, column=3).value == "Kwork"
    wb.close()
